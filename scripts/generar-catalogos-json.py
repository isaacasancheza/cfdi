#!/usr/bin/env -S uv run
import json
from argparse import ArgumentParser, ArgumentTypeError
from functools import cache
from pathlib import Path
from typing import Annotated, Any, Generic, TypeVar

from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, BeforeValidator, StringConstraints


def convert_any_to_string(
    value: Any,
):
    if not isinstance(value, str):
        return str(value)
    return value


def format_int_as_02d(
    value: Any,
):
    if isinstance(value, int):
        return f'{int(value):02d}'
    return value


def format_int_as_03d(
    value: Any,
):
    if isinstance(value, int):
        return f'{int(value):03d}'
    return value


type Str = Annotated[
    str,
    BeforeValidator(convert_any_to_string),
    StringConstraints(
        min_length=1,
        strip_whitespace=True,
    ),
]
type Int02 = Annotated[str, BeforeValidator(format_int_as_02d)]
type Int03 = Annotated[str, BeforeValidator(format_int_as_03d)]

Key = TypeVar('Key', Str, Int02, Int03)


class KeyDescriptionModel(BaseModel, Generic[Key]):
    clave: Key
    descripcion: Annotated[
        str,
        StringConstraints(
            min_length=1,
            strip_whitespace=True,
        ),
    ]


class StrModel(KeyDescriptionModel[Str]):
    pass


class Int02Model(KeyDescriptionModel[Int02]):
    pass


class Int03Model(KeyDescriptionModel[Int03]):
    pass


Model = TypeVar('Model', StrModel, Int02Model, Int03Model)

Catalogo = dict[str, dict[str, str]]


@cache
def get_workbook() -> Workbook:
    return load_workbook(filename='../assets/catalogos.xlsx')


def get_catalogo(
    *,
    min_row: int,
    max_row: int,
    model: type[Model],
    worksheet: str,
):
    wb = get_workbook()
    catalogo: Catalogo = {}
    for row in wb[worksheet].iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=1,
        max_col=2,
    ):
        clave, descripcion = row
        if clave.value is None:
            continue
        assert isinstance(descripcion.value, str)
        entry = model.model_validate(
            {
                'clave': clave.value,
                'descripcion': descripcion.value.rstrip('.'),
            }
        )
        entry = entry.model_dump()
        catalogo[entry['clave']] = entry['descripcion']
    return catalogo


def get_forma_pago():
    return get_catalogo(
        min_row=7,
        max_row=28,
        model=Int02Model,
        worksheet='c_FormaPago',
    )


def get_moneda():
    return get_catalogo(
        min_row=6,
        max_row=188,
        model=StrModel,
        worksheet='c_Moneda',
    )


def get_tipo_de_comprobante():
    return get_catalogo(
        min_row=6,
        max_row=11,
        model=StrModel,
        worksheet='c_TipoDeComprobante',
    )


def get_exportacion():
    return get_catalogo(
        min_row=6,
        max_row=9,
        model=Int02Model,
        worksheet='c_Exportacion',
    )


def get_metodo_pago():
    return get_catalogo(
        min_row=7,
        max_row=8,
        model=StrModel,
        worksheet='c_MetodoPago',
    )


def get_peridiocidad():
    return get_catalogo(
        min_row=6,
        max_row=10,
        model=Int02Model,
        worksheet='c_Periodicidad',
    )


def get_meses():
    return get_catalogo(
        min_row=6,
        max_row=23,
        model=Int02Model,
        worksheet='c_Meses',
    )


def get_tipo_relacion():
    return get_catalogo(
        min_row=6,
        max_row=12,
        model=Int02Model,
        worksheet='c_TipoRelacion',
    )


def get_regimen_fiscal():
    return get_catalogo(
        min_row=7,
        max_row=25,
        model=Int03Model,
        worksheet='c_RegimenFiscal',
    )


def get_pais():
    return get_catalogo(
        min_row=6,
        max_row=255,
        model=StrModel,
        worksheet='c_Pais',
    )


def get_uso_cfdi():
    return get_catalogo(
        min_row=7,
        max_row=30,
        model=StrModel,
        worksheet='c_UsoCFDI',
    )


def get_clave_prod_serv():
    return get_catalogo(
        min_row=6,
        max_row=52518,
        model=StrModel,
        worksheet='c_ClaveProdServ',
    )


def get_clave_unidad():
    return get_catalogo(
        min_row=7,
        max_row=2424,
        model=StrModel,
        worksheet='c_ClaveUnidad',
    )


def get_objeto_imp():
    return get_catalogo(
        min_row=6,
        max_row=13,
        model=Int02Model,
        worksheet='c_ObjetoImp',
    )


def get_impuesto():
    return get_catalogo(
        min_row=6,
        max_row=8,
        model=Int03Model,
        worksheet='c_Impuesto',
    )


def get_aduana():
    return get_catalogo(
        min_row=6,
        max_row=55,
        model=Int02Model,
        worksheet='c_Aduana',
    )


def validate_directory(
    raw_path: str,
):
    path = Path(raw_path)
    if not path.is_dir():
        raise ArgumentTypeError(
            f'the directory {path} does not exist or is not a valid directory.'
        )
    return path


if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument('output', type=validate_directory)

    args = parser.parse_args()
    output: Path = args.output

    forma_pago = get_forma_pago()
    moneda = get_moneda()
    tipo_de_comprobante = get_tipo_de_comprobante()
    exportacion = get_exportacion()
    metodo_pago = get_metodo_pago()
    peridiocidad = get_peridiocidad()
    meses = get_meses()
    tipo_relacion = get_tipo_relacion()
    regimen_fiscal = get_regimen_fiscal()
    pais = get_pais()
    uso_cfdi = get_uso_cfdi()
    clave_prod_serv = get_clave_prod_serv()
    clave_unidad = get_clave_unidad()
    objeto_imp = get_objeto_imp()
    impuesto = get_impuesto()
    aduana = get_aduana()

    catalogos = {
        'forma_pago': forma_pago,
        'moneda': moneda,
        'tipo_de_comprobante': tipo_de_comprobante,
        'exportacion': exportacion,
        'metodo_pago': metodo_pago,
        'peridiocidad': peridiocidad,
        'meses': meses,
        'tipo_relacion': tipo_relacion,
        'regimen_fiscal': regimen_fiscal,
        'pais': pais,
        'uso_cfdi': uso_cfdi,
        'clave_prod_serv': clave_prod_serv,
        'clave_unidad': clave_unidad,
        'objeto_imp': objeto_imp,
        'impuesto': impuesto,
        'aduana': aduana,
    }
    for catalogo in catalogos:
        with open(output / f'{catalogo}.json', 'w') as f:
            json.dump(catalogos[catalogo], f, indent=4)
