import ast
from argparse import ArgumentParser, ArgumentTypeError
from pathlib import Path
from typing import Any, Callable, Generator, TypedDict

import astor
from boto3 import resource
from openpyxl import Workbook, load_workbook

from cfdi.v40.pdf import Catalogo

dynamodb = resource('dynamodb')


class CreateEnum(TypedDict):
    min_row: int
    worksheet: str
    module_name: str
    enum_name: str
    enum_values_prefix: str
    enum_value_as_name: bool
    formatter: Callable[[Any], str | None] | None
    single_column: bool


class CreateDatabase(TypedDict):
    min_row: int
    worksheet: str
    formatter: Callable[[Any], str | None] | None
    keys: tuple[str, ...]
    pk: Catalogo


def int02d(
    value: str | None,
    /,
):
    if value is None:
        return
    return f'{int(value):02d}'


def int03d(
    value: str | None,
    /,
):
    if value is None:
        return
    return f'{int(value):03d}'


def int04d(
    value: str | None,
    /,
):
    if value is None:
        return
    return f'{int(value):04d}'


def int05d(
    value: str | None,
    /,
):
    if value is None:
        return
    return f'{int(value):05d}'


ENUMS: list[CreateEnum] = [
    # forma pago
    {
        'module_name': 'forma_pago',
        'enum_name': 'FormaPago',
        'enum_value_as_name': False,
        'enum_values_prefix': 'FP_',
        'formatter': int02d,
        'min_row': 7,
        'worksheet': 'c_FormaPago',
        'single_column': False,
    },
    # moneda
    {
        'module_name': 'moneda',
        'enum_name': 'Moneda',
        'enum_value_as_name': False,
        'enum_values_prefix': 'M_',
        'formatter': None,
        'min_row': 6,
        'worksheet': 'c_Moneda',
        'single_column': False,
    },
    # tipo de comprobante
    {
        'module_name': 'tipo_de_comprobante',
        'enum_name': 'TipoDeComprobante',
        'enum_value_as_name': True,
        'enum_values_prefix': 'TC_',
        'formatter': None,
        'min_row': 6,
        'worksheet': 'c_TipoDeComprobante',
        'single_column': False,
    },
    # exportacion
    {
        'module_name': 'exportacion',
        'enum_name': 'Exportacion',
        'enum_value_as_name': False,
        'enum_values_prefix': 'E_',
        'formatter': int02d,
        'min_row': 6,
        'worksheet': 'c_Exportacion',
        'single_column': False,
    },
    # metodo pago
    {
        'module_name': 'metodo_pago',
        'enum_name': 'MetodoPago',
        'enum_value_as_name': False,
        'enum_values_prefix': 'MP_',
        'formatter': None,
        'min_row': 7,
        'worksheet': 'c_MetodoPago',
        'single_column': False,
    },
    # Periodicidad
    {
        'module_name': 'periodicidad',
        'enum_name': 'Periodicidad',
        'enum_value_as_name': True,
        'enum_values_prefix': 'P_',
        'formatter': int02d,
        'min_row': 7,
        'worksheet': 'c_Periodicidad',
        'single_column': False,
    },
    # meses
    {
        'module_name': 'meses',
        'enum_name': 'Meses',
        'enum_value_as_name': False,
        'enum_values_prefix': 'M_',
        'formatter': int02d,
        'min_row': 6,
        'worksheet': 'c_Meses',
        'single_column': False,
    },
    # tipo relacion
    {
        'module_name': 'tipo_relacion',
        'enum_name': 'TipoRelacion',
        'enum_value_as_name': False,
        'enum_values_prefix': 'TR_',
        'formatter': int02d,
        'min_row': 6,
        'worksheet': 'c_TipoRelacion',
        'single_column': False,
    },
    # regimen fiscal
    {
        'module_name': 'regimen_fiscal',
        'enum_name': 'RegimenFiscal',
        'enum_value_as_name': False,
        'enum_values_prefix': 'RF_',
        'formatter': None,
        'min_row': 7,
        'worksheet': 'c_RegimenFiscal',
        'single_column': False,
    },
    # pais
    {
        'module_name': 'pais',
        'enum_name': 'Pais',
        'enum_value_as_name': False,
        'enum_values_prefix': 'P_',
        'formatter': None,
        'min_row': 6,
        'worksheet': 'c_Pais',
        'single_column': False,
    },
    # uso cfdi
    {
        'module_name': 'uso_cfdi',
        'enum_name': 'UsoCFDI',
        'enum_value_as_name': False,
        'enum_values_prefix': 'UC_',
        'formatter': None,
        'min_row': 7,
        'worksheet': 'c_UsoCFDI',
        'single_column': False,
    },
    # too big
    # # clave prod serv
    # {
    #     'module_name': 'clave_prod_serv',
    #     'enum_name': 'ClaveProdServ',
    #     'enum_value_as_name': False,
    #     'enum_values_prefix': 'CPS_',
    #     'formatter': None,
    #     'min_row': 6,
    #     'worksheet': 'c_ClaveProdServ',
    #     'single_column': False,
    # },
    # clave unidad
    {
        'module_name': 'clave_unidad',
        'enum_name': 'ClaveUnidad',
        'enum_value_as_name': False,
        'enum_values_prefix': 'CU_',
        'formatter': None,
        'min_row': 7,
        'worksheet': 'c_ClaveUnidad',
        'single_column': False,
    },
    # objeto imp
    {
        'module_name': 'objeto_imp',
        'enum_name': 'ObjetoImp',
        'enum_value_as_name': False,
        'enum_values_prefix': 'OI_',
        'formatter': int02d,
        'min_row': 6,
        'worksheet': 'c_ObjetoImp',
        'single_column': False,
    },
    # impuesto
    {
        'module_name': 'impuesto',
        'enum_name': 'Impuesto',
        'enum_value_as_name': False,
        'enum_values_prefix': 'I_',
        'formatter': int03d,
        'min_row': 6,
        'worksheet': 'c_Impuesto',
        'single_column': False,
    },
    # aduana
    {
        'module_name': 'aduana',
        'enum_name': 'Aduana',
        'enum_value_as_name': False,
        'enum_values_prefix': 'A_',
        'formatter': int02d,
        'min_row': 6,
        'worksheet': 'c_Aduana',
        'single_column': False,
    },
    # tipo factor
    {
        'module_name': 'tipo_factor',
        'enum_name': 'TipoFactor',
        'enum_value_as_name': True,
        'enum_values_prefix': 'TF_',
        'formatter': None,
        'min_row': 6,
        'worksheet': 'c_TipoFactor',
        'single_column': True,
    },
]


DATABASE: list[CreateDatabase] = [
    # forma pago
    {
        'min_row': 7,
        'worksheet': 'c_FormaPago',
        'formatter': int02d,
        'keys': ('sk', 'name'),
        'pk': 'FORMA_PAGO',
    },
    # moneda
    {
        'min_row': 6,
        'worksheet': 'c_Moneda',
        'formatter': None,
        'keys': ('sk', 'name'),
        'pk': 'MONEDA',
    },
    # tipo de comprobante
    {
        'min_row': 6,
        'worksheet': 'c_TipoDeComprobante',
        'formatter': None,
        'keys': ('sk', 'name'),
        'pk': 'TIPO_DE_COMPROBANTE',
    },
    # exportacion
    {
        'min_row': 6,
        'worksheet': 'c_Exportacion',
        'formatter': int02d,
        'keys': ('sk', 'name'),
        'pk': 'EXPORTACION',
    },
    #  metodo pago
    {
        'min_row': 7,
        'worksheet': 'c_MetodoPago',
        'formatter': None,
        'keys': ('sk', 'name'),
        'pk': 'METODO_PAGO',
    },
    # periodicidad
    {
        'min_row': 7,
        'worksheet': 'c_Periodicidad',
        'formatter': int02d,
        'keys': ('sk', 'name'),
        'pk': 'PERIODICIDAD',
    },
    # meses
    {
        'min_row': 6,
        'worksheet': 'c_Meses',
        'formatter': int02d,
        'keys': ('sk', 'name'),
        'pk': 'MESES',
    },
    # tipo relacion
    {
        'min_row': 6,
        'worksheet': 'c_TipoRelacion',
        'formatter': int02d,
        'keys': ('sk', 'name'),
        'pk': 'TIPO_RELACION',
    },
    # regimen fiscal
    {
        'min_row': 7,
        'worksheet': 'c_RegimenFiscal',
        'formatter': None,
        'keys': ('sk', 'name'),
        'pk': 'REGIMEN_FISCAL',
    },
    # pais
    {
        'min_row': 6,
        'worksheet': 'c_Pais',
        'formatter': None,
        'keys': ('sk', 'name'),
        'pk': 'PAIS',
    },
    # uso cfdi
    {
        'min_row': 7,
        'worksheet': 'c_UsoCFDI',
        'formatter': None,
        'keys': ('sk', 'name'),
        'pk': 'USO_CFDI',
    },
    # clave prod serv
    {
        'min_row': 6,
        'worksheet': 'c_ClaveProdServ',
        'formatter': None,
        'keys': ('sk', 'name'),
        'pk': 'CLAVE_PROD_SERV',
    },
    # clave unidad
    {
        'min_row': 7,
        'worksheet': 'c_ClaveUnidad',
        'formatter': None,
        'keys': ('sk', 'name'),
        'pk': 'CLAVE_UNIDAD',
    },
    # objeto imp
    {
        'min_row': 6,
        'worksheet': 'c_ObjetoImp',
        'formatter': int02d,
        'keys': ('sk', 'name'),
        'pk': 'OBJETO_IMP',
    },
    # impuesto
    {
        'min_row': 6,
        'worksheet': 'c_Impuesto',
        'formatter': int03d,
        'keys': ('sk', 'name'),
        'pk': 'IMPUESTO',
    },
    # aduana
    {
        'min_row': 6,
        'worksheet': 'c_Aduana',
        'formatter': int02d,
        'keys': ('sk', 'name'),
        'pk': 'ADUANA',
    },
    # tipo factor
    {
        'min_row': 6,
        'worksheet': 'c_TipoFactor',
        'formatter': None,
        'keys': ('sk', 'name'),
        'pk': 'TIPO_FACTOR',
    },
]


def extract_data(
    *,
    min_row: int,
    max_col: int,
    workbook: Workbook,
    worksheet: str,
) -> Generator[Generator[str | None]]:
    for row in workbook[worksheet].iter_rows(
        min_row=min_row,
        max_col=max_col,
    ):
        if row[0].value is None:
            continue
        yield (
            str(cell.value).strip() if cell.value is not None else None for cell in row
        )


def create_enum(
    *,
    enum_name: str,
    enum_values: Generator[Generator[str | None]],
    formatter: Callable[[Any], str | None] | None = None,
    single_column: bool = False,
    enum_values_prefix: str,
    enum_value_as_name: bool = False,
):
    body = []
    for pair in enum_values:
        if single_column:
            v = next(pair)
            key = v
            value = v
        else:
            key, value = pair
        if formatter is not None:
            key = formatter(key)
        if value is None:
            continue
        name = key
        if enum_value_as_name:
            name = value.upper()
        if enum_values_prefix is not None:
            name = f'{enum_values_prefix}{name}'
        body.append(
            ast.Assign(
                targets=[
                    ast.Name(
                        id=name,
                        ctx=ast.Store(),
                    ),
                ],
                value=ast.Constant(
                    value=key,
                ),
            )
        )
        body.append(
            ast.Expr(
                value=ast.Constant(
                    value=value,
                ),
            ),
        )
    enum = ast.ClassDef(
        name=enum_name,
        bases=[
            ast.Name(
                id='StrEnum',
                ctx=ast.Load(),
            ),
        ],
        body=body,
    )

    # import enum.StreEnum
    import_enum = ast.ImportFrom(
        module='enum',
        names=[
            ast.alias(
                name='StrEnum',
            ),
        ],
        level=0,
    )

    # create module
    module = ast.Module(
        body=[
            import_enum,
            enum,
        ],
    )

    # convert ast to source code
    source_code = astor.to_source(module)

    return source_code


def create_enums(
    *,
    output_dir: Path,
    workbook_file: Path,
) -> None:
    package_root = output_dir / 'sat'
    catalogos_module = package_root / 'catalogos'

    output_dir.mkdir(exist_ok=True)
    catalogos_module.mkdir(exist_ok=True)

    package_init = output_dir / '__init__.py'
    package_init.touch(exist_ok=True)

    workbook = load_workbook(workbook_file)
    enums: list[str] = []
    imports: list[ast.stmt] = []
    for enum in ENUMS:
        enum_values = extract_data(
            min_row=enum['min_row'],
            max_col=1 if enum['single_column'] else 2,
            workbook=workbook,
            worksheet=enum['worksheet'],
        )

        source_code = create_enum(
            enum_name=enum['enum_name'],
            enum_values=enum_values,
            formatter=enum['formatter'],
            single_column=enum['single_column'],
            enum_values_prefix=enum['enum_values_prefix'],
            enum_value_as_name=enum['enum_value_as_name'],
        )

        imprt = ast.ImportFrom(
            module=enum['module_name'],
            names=[
                ast.alias(
                    name=enum['enum_name'],
                ),
            ],
            level=1,
        )

        module_path = catalogos_module / f'{enum["module_name"]}.py'
        with open(module_path, 'w') as f:
            f.write(source_code)

        enums.append(enum['enum_name'])
        imports.append(imprt)

    # import enums
    init_module = ast.Module(body=imports)
    init_source_code = astor.to_source(init_module)
    with open(catalogos_module / '__init__.py', 'w', encoding='utf-8') as f:
        f.write(init_source_code)


def extract_codigos_postales(
    workbook: Workbook,
):
    for i in range(1, 3):
        yield extract_data(
            min_row=8,
            max_col=4,
            workbook=workbook,
            worksheet=f'c_CodigoPostal_Parte_{i}',
        )


def extract_colonias(
    workbook: Workbook,
):
    for i in range(1, 4):
        yield extract_data(
            min_row=6,
            max_col=3,
            workbook=workbook,
            worksheet=f'C_Colonia_{i}',
        )


def extract_estados(
    workbook: Workbook,
):
    yield extract_data(
        min_row=6,
        max_col=3,
        workbook=workbook,
        worksheet='c_Estado',
    )


def extract_localidades(
    workbook: Workbook,
):
    yield extract_data(
        min_row=6,
        max_col=3,
        workbook=workbook,
        worksheet='c_Localidad',
    )


def extract_municipios(
    workbook: Workbook,
):
    yield extract_data(
        min_row=6,
        max_col=3,
        workbook=workbook,
        worksheet='c_Municipio',
    )


def create_database(
    workbook_file: Path,
    table_name: str,
):
    table = dynamodb.Table(table_name)
    workbook = load_workbook(workbook_file)

    for entry in DATABASE:
        values = extract_data(
            min_row=entry['min_row'],
            max_col=2,
            workbook=workbook,
            worksheet=entry['worksheet'],
        )
        data = (dict(zip(entry['keys'], v)) for v in values)
        with table.batch_writer() as batch:
            for item in data:
                item['pk'] = entry['pk']
                if entry['formatter'] is not None:
                    item['sk'] = entry['formatter'](item['sk'])
                batch.put_item(Item=item)

    # too heavy
    # postal_code_keys = (
    #     'sk',
    #     'estado_sk',
    #     'municipio_sk',
    #     'localidad_sk',
    # )

    # for entry in extract_codigos_postales(workbook):
    #     data = (dict(zip(postal_code_keys, v)) for v in entry)
    #     with table.batch_writer() as batch:
    #         for item in data:
    #             item['sk'] = int05d(item['sk'])
    #             item['municipio_sk'] = int03d(item['municipio_sk'])
    #             item['localidad_sk'] = int02d(item['localidad_sk'])
    #             item = cast(dict, item)
    #             item['pk'] = 'CODIGO_POSTAL'
    #             batch.put_item(Item=item)

    # unnecessary
    # colonias_keys = (
    #     'sk',
    #     'codigo_postal_sk',
    #     'name',
    # )
    # for entry in extract_colonias(workbook):
    #     data = (dict(zip(colonias_keys, v)) for v in entry)
    #     with table.batch_writer() as batch:
    #         for item in data:
    #             item['sk'] = int04d(item['sk'])
    #             item['codigo_postal_sk'] = int05d(item['codigo_postal_sk'])
    #             item = cast(dict, item)
    #             item['pk'] = 'COLONIA'
    #             batch.put_item(Item=item)

    # estados_keys = [
    #     'sk',
    #     'pais_sk',
    #     'name',
    # ]
    # for entry in extract_estados(workbook):
    #     data = (dict(zip(estados_keys, v)) for v in entry)
    #     with table.batch_writer() as batch:
    #         for item in data:
    #             item = cast(dict, item)
    #             item['pk'] = 'ESTADO'
    #             batch.put_item(Item=item)

    # localidates_keys = [
    #     'sk',
    #     'estado_sk',
    #     'name',
    # ]
    # for entry in extract_localidades(workbook):
    #     data = (dict(zip(localidates_keys, v)) for v in entry)
    #     with table.batch_writer() as batch:
    #         for item in data:
    #             item['sk'] = int02d(item['sk'])
    #             item = cast(dict, item)
    #             item['pk'] = 'LOCALIDAD'
    #             batch.put_item(Item=item)

    # municipios_keys = [
    #     'sk',
    #     'estado_sk',
    #     'name',
    # ]
    # for entry in extract_municipios(workbook):
    #     data = (dict(zip(municipios_keys, v)) for v in entry)
    #     with table.batch_writer() as batch:
    #         for item in data:
    #             item['sk'] = int03d(item['sk'])
    #             item = cast(dict, item)
    #             item['pk'] = 'MUNICIPIO'
    #             batch.put_item(Item=item)


def file(
    filename: str,
) -> Path:
    path = Path(filename)
    if not path.exists():
        raise ArgumentTypeError(f'file {filename} not found')
    return path


def directory(
    dirname: str,
) -> Path:
    path = Path(dirname)
    if not path.is_dir():
        raise ArgumentTypeError(f'directory {dirname} not found')
    return path


if __name__ == '__main__':
    parser = ArgumentParser()

    action = parser.add_subparsers(dest='action', required=True)

    enums_parser = action.add_parser('generate-enums')
    enums_parser.add_argument('-w', '--workbook', type=file, required=True)
    enums_parser.add_argument('-o', '--output', type=directory, required=True)

    dynamodb_parser = action.add_parser('generate-database')
    dynamodb_parser.add_argument('-w', '--workbook', type=file, required=True)
    dynamodb_parser.add_argument('-t', '--table-name', type=str, required=True)

    args = parser.parse_args()

    match args.action:
        case 'generate-enums':
            create_enums(
                workbook_file=args.workbook,
                output_dir=args.output,
            )
        case 'generate-database':
            create_database(
                workbook_file=args.workbook,
                table_name=args.table_name,
            )
